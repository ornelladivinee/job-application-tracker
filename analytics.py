import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

# =========================
# ARCHIVO
# =========================
file_path = "applications_tracker.xlsx"

# =========================
# LEER DATOS
# =========================
df = pd.read_excel(file_path)

# =========================
# CALCULAR ESTADÍSTICAS
# =========================
total = len(df)

applied = len(df[df["Status"] == "Applied"])
rejected = len(df[df["Status"] == "Rejected"])
pending = len(df[df["Status"] == "Pending"])

# =========================
# ABRIR EXCEL
# =========================
wb = load_workbook(file_path)
ws = wb.active

# =========================
# ELIMINAR RESUMEN ANTERIOR
# =========================
for row in range(1, ws.max_row + 1):
    cell_value = ws.cell(row=row, column=1).value

    if cell_value == "JOB TRACKER DASHBOARD":
        ws.delete_rows(row, 6)
        break

# =========================
# NUEVA POSICIÓN DEL RESUMEN
# =========================
summary_row = ws.max_row + 3

# =========================
# ESCRIBIR DASHBOARD
# =========================
title_cell = ws.cell(row=summary_row, column=1)
title_cell.value = "JOB TRACKER DASHBOARD"
title_cell.font = Font(bold=True, size=14)

ws.cell(row=summary_row + 1, column=1, value="Total Applications")
ws.cell(row=summary_row + 1, column=2, value=total)

ws.cell(row=summary_row + 2, column=1, value="Applied")
ws.cell(row=summary_row + 2, column=2, value=applied)

ws.cell(row=summary_row + 3, column=1, value="Rejected")
ws.cell(row=summary_row + 3, column=2, value=rejected)

ws.cell(row=summary_row + 4, column=1, value="Pending")
ws.cell(row=summary_row + 4, column=2, value=pending)

# =========================
# GUARDAR
# =========================
wb.save(file_path)

print("✔ Dashboard actualizado correctamente")