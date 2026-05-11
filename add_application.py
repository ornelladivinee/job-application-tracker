import pandas as pd
from datetime import datetime

# Cargar Excel existente
df = pd.read_excel("applications_tracker.xlsx")

# Lista de postulaciones
applications = [
    {
        "Company": "NTT Data",
        "Country": "Spain - A Coruña",
        "Position": "Prácticas Ingeniería Informática - Telecomunicaciones",
        "Date Applied": "2026-05-06",
        "Platform": "iAgora",
        "Status": "Applied",
        "Link": "",
        "Notes": "IT internship"
    },
    {
        "Company": "Airbus",
        "Country": "Spain - Getafe",
        "Position": "#Discover II 2026-2027 Internship FDT Stress Team",
        "Date Applied": "2026-05-06",
        "Platform": "iAgora",
        "Status": "Applied",
        "Link": "",
        "Notes": "Engineering / Stress team"
    },
    {
        "Company": "Siemens",
        "Country": "Portugal - Porto",
        "Position": "Azure DevOps Trainee (m/f/d)",
        "Date Applied": "2026-05-06",
        "Platform": "iAgora",
        "Status": "Rejected",
        "Link": "",
        "Notes": "Cloud / DevOps / Azure"
    },
    {
        "Company": "Airbus",
        "Country": "Portugal - Lisbon",
        "Position": "IT Trainee",
        "Date Applied": "2026-05-06",
        "Platform": "iAgora",
        "Status": "Rejected",
        "Link": "",
        "Notes": "IT Systems"
    },
    {
        "Company": "Capgemini",
        "Country": "Spain - Valencia",
        "Position": "Prácticas en Ciberseguridad",
        "Date Applied": "",
        "Platform": "iAgora",
        "Status": "Applied",
        "Link": "",
        "Notes": "Cybersecurity internship"
    },
    {
        "Company": "Bosch",
        "Country": "Portugal",
        "Position": "Data Engineering Internship (Lakehouse / Databricks)",
        "Date Applied": "2026-05-07",
        "Platform": "iAgora",
        "Status": "Applied",
        "Link": "",
        "Notes": "Data engineering, cloud, Hadoop to Databricks migration"
    },
    {
        "Company": "Zurich Insurance",
        "Country": "Spain",
        "Position": "Trainee Cloud Application Developer - 133025",
        "Date Applied": "2026-05-07",
        "Platform": "Zurich Careers / iAgora",
        "Status": "Rejected",
        "Focus": "Cloud Development",
        "Notes": "Application rejected after recruitment process review"
    },
    {
        "Company": "Nokia",
        "Country": "Portugal",
        "City": "Portugal (location not specified)",
        "Position": "Telecommunication Support Trainee - Global Core Network Team",
        "Date Applied": "2026-05-07",
        "Platform": "iAgora / Job Portal",
        "Status": "Applied",
        "Work Mode": "International / Hybrid",
        "Technologies": [
            "IMS Voice",
            "Packet Core",
            "Cloud Infrastructure",
            "Linux",
            "OpenStack",
            "3GPP",
            "VoLTE / VoWiFi / Vo5G"
        ],
        "Focus Area": "Telecommunications / Cloud / Network Engineering",
        "Notes": "Strong technical telecom role with exposure to cloud and network core systems"
    },
    {
        "Company": "Siemens",
        "Country": "Romania",
        "City": "Brasov / Cluj-Napoca / Bucharest",
        "Position": "Siemens Talents Hub Internship 2026 - Data Analytics & Artificial Intelligence",
        "Date Applied": "2026-05-07",
        "Platform": "iAgora / Siemens Careers",
        "Status": "Applied",
        "Focus": "AI, Data Analytics, R&D, Machine Learning, Software Prototyping",
        "Duration": "3 months",
        "Paid": "No",
        "Notes": "R&D internship in AI and Data Analytics"
    },
    {
        "Company": "Novo Nordisk",
        "Country": "Denmark",
        "City": "Bagsværd",
        "Position": "BI, Data & AI Intern - Complaint Insights Team",
        "Date Applied": "2026-05-07",
        "Platform": "Career Portal / Application Form",
        "Status": "Rejected",
        "Focus": "Business Intelligence, Data Analytics, AI, Power BI",
        "Tools": [
            "Power BI",
            "Python",
            "SQL",
            "AI tools"
        ],
        "Duration": "Internship",
        "Paid": "Yes",
        "Notes": "BI + AI internship focused on healthcare data analytics"
    },
    {
    "Company": "NTT DATA",
    "Country": "Spain",
    "City": "A Coruña",
    "Position": "Prácticas Ingeniería Informática - Telecomunicaciones",
    "Date Applied": "2026-05-10",
    "Platform": "NTT DATA Careers / iAgora",
    "Status": "Applied",
    "Focus": "Software Development, Telecommunications, IT",
    "Paid": "Not specified",
    "Notes": "Application confirmed by NTT DATA recruitment team via email"
},

{
    "Company": "BBVA",
    "Country": "Spain",
    "City": "Madrid",
    "Position": "BECAS BBVA ÁREAS DE TECNOLOGÍA E INNOVACIÓN 2026",
    "Date Applied": "2026-05-10",
    "Platform": "BBVA Careers",
    "Status": "Applied",
    "Focus": "Technology, Innovation, IT",
    "Paid": "Not specified",
    "Notes": "Scholarship/internship application confirmed by BBVA recruitment team"
},
{
    "Company": "Natixis",
    "Country": "Portugal",
    "City": "Porto / Lisbon",
    "Position": "IT Trainee | WE WANT YOU 2026 S2",
    "Date Applied": "2026-05-11",
    "Platform": "Natixis Careers",
    "Status": "Applied",
    "Focus": "IT Trainee Program, Development, Data, Cybersecurity, DevOps",
    "Paid": "Not specified",
    "Work Mode": "Hybrid",
    "Notes": "International trainee program focused on IT, software development, data, cybersecurity and analytics"
}
]
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

# =========================
# LISTA DE POSTULACIONES (TUS 13)
# =========================
applications = [
    {
        "Company": "NTT Data",
        "Country": "Spain - A Coruña",
        "Position": "Prácticas Ingeniería Informática - Telecomunicaciones",
        "Date Applied": "2026-05-06",
        "Platform": "iAgora",
        "Status": "Applied",
        "Link": "",
        "Notes": "IT internship"
    },
    {
        "Company": "Airbus",
        "Country": "Spain - Getafe",
        "Position": "#Discover II 2026-2027 Internship FDT Stress Team",
        "Date Applied": "2026-05-06",
        "Platform": "iAgora",
        "Status": "Applied",
        "Link": "",
        "Notes": "Engineering / Stress team"
    },
    {
        "Company": "Siemens",
        "Country": "Portugal - Porto",
        "Position": "Azure DevOps Trainee (m/f/d)",
        "Date Applied": "2026-05-06",
        "Platform": "iAgora",
        "Status": "Rejected",
        "Link": "",
        "Notes": "Cloud / DevOps / Azure"
    },
    {
        "Company": "Airbus",
        "Country": "Portugal - Lisbon",
        "Position": "IT Trainee",
        "Date Applied": "2026-05-06",
        "Platform": "iAgora",
        "Status": "Rejected",
        "Link": "",
        "Notes": "IT Systems"
    },
    {
        "Company": "Capgemini",
        "Country": "Spain - Valencia",
        "Position": "Prácticas en Ciberseguridad",
        "Date Applied": "",
        "Platform": "iAgora",
        "Status": "Applied",
        "Link": "",
        "Notes": "Cybersecurity internship"
    },
    {
        "Company": "Bosch",
        "Country": "Portugal",
        "Position": "Data Engineering Internship (Lakehouse / Databricks)",
        "Date Applied": "2026-05-07",
        "Platform": "iAgora",
        "Status": "Applied",
        "Link": "",
        "Notes": "Data engineering, cloud, Hadoop to Databricks migration"
    },
    {
        "Company": "Zurich Insurance",
        "Country": "Spain",
        "Position": "Trainee Cloud Application Developer - 133025",
        "Date Applied": "2026-05-07",
        "Platform": "Zurich Careers / iAgora",
        "Status": "Rejected",
        "Focus": "Cloud Development",
        "Notes": "Application rejected after recruitment process review"
    },
    {
        "Company": "Nokia",
        "Country": "Portugal",
        "City": "Portugal (location not specified)",
        "Position": "Telecommunication Support Trainee - Global Core Network Team",
        "Date Applied": "2026-05-07",
        "Platform": "iAgora / Job Portal",
        "Status": "Applied",
        "Work Mode": "International / Hybrid",
        "Technologies": [
            "IMS Voice",
            "Packet Core",
            "Cloud Infrastructure",
            "Linux",
            "OpenStack",
            "3GPP",
            "VoLTE / VoWiFi / Vo5G"
        ],
        "Focus Area": "Telecommunications / Cloud / Network Engineering",
        "Notes": "Strong technical telecom role with exposure to cloud and network core systems"
    },
    {
        "Company": "Siemens",
        "Country": "Romania",
        "City": "Brasov / Cluj-Napoca / Bucharest",
        "Position": "Siemens Talents Hub Internship 2026 - Data Analytics & Artificial Intelligence",
        "Date Applied": "2026-05-07",
        "Platform": "iAgora / Siemens Careers",
        "Status": "Applied",
        "Focus": "AI, Data Analytics, R&D, Machine Learning, Software Prototyping",
        "Duration": "3 months",
        "Paid": "No",
        "Notes": "R&D internship in AI and Data Analytics"
    },
    {
        "Company": "Novo Nordisk",
        "Country": "Denmark",
        "City": "Bagsværd",
        "Position": "BI, Data & AI Intern - Complaint Insights Team",
        "Date Applied": "2026-05-07",
        "Platform": "Career Portal / Application Form",
        "Status": "Rejected",
        "Focus": "Business Intelligence, Data Analytics, AI, Power BI",
        "Tools": [
            "Power BI",
            "Python",
            "SQL",
            "AI tools"
        ],
        "Duration": "Internship",
        "Paid": "Yes",
        "Notes": "BI + AI internship focused on healthcare data analytics"
    },
    {
        "Company": "NTT DATA",
        "Country": "Spain",
        "City": "A Coruña",
        "Position": "Prácticas Ingeniería Informática - Telecomunicaciones",
        "Date Applied": "2026-05-10",
        "Platform": "NTT DATA Careers / iAgora",
        "Status": "Applied",
        "Focus": "Software Development, Telecommunications, IT",
        "Paid": "Not specified",
        "Notes": "Application confirmed by NTT DATA recruitment team via email"
    },
    {
        "Company": "BBVA",
        "Country": "Spain",
        "City": "Madrid",
        "Position": "BECAS BBVA ÁREAS DE TECNOLOGÍA E INNOVACIÓN 2026",
        "Date Applied": "2026-05-10",
        "Platform": "BBVA Careers",
        "Status": "Applied",
        "Focus": "Technology, Innovation, IT",
        "Paid": "Not specified",
        "Notes": "Scholarship/internship application confirmed by BBVA recruitment team"
    },
    {
        "Company": "Natixis",
        "Country": "Portugal",
        "City": "Porto / Lisbon",
        "Position": "IT Trainee | WE WANT YOU 2026 S2",
        "Date Applied": "2026-05-11",
        "Platform": "Natixis Careers",
        "Status": "Applied",
        "Focus": "IT Trainee Program, Development, Data, Cybersecurity, DevOps",
        "Paid": "Not specified",
        "Work Mode": "Hybrid",
        "Notes": "International trainee program focused on IT, software development, data, cybersecurity and analytics"
    }
]

# =========================
# CREAR DATAFRAME Y SOBRESCRIBIR EL EXCEL (SIN ACUMULAR)
# =========================
df = pd.DataFrame(applications)
df.to_excel("applications_tracker.xlsx", index=False)
print(f"Archivo sobrescrito con {len(df)} postulaciones.")

# =========================
# FORMATO CON OPENPYXL
# =========================
wb = load_workbook("applications_tracker.xlsx")
ws = wb.active

# --- Limpiar columnas ID anteriores y crear nueva ---
for col_idx in range(ws.max_column, 0, -1):
    header = ws.cell(row=1, column=col_idx).value
    if header and str(header).startswith("ID"):
        ws.delete_cols(col_idx)

ws.insert_cols(1)
ws["A1"] = "ID"
for row in range(2, ws.max_row + 1):
    ws[f"A{row}"] = row - 1

# --- Estilos de encabezado ---
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font

# --- Colores por estado ---
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

# Localizar columna Status
status_col = None
for cell in ws[1]:
    if cell.value == "Status":
        status_col = cell.column
        break

if status_col:
    for row in range(2, ws.max_row + 1):
        status_cell = ws.cell(row=row, column=status_col)
        if status_cell.value == "Applied":
            status_cell.fill = green_fill
        elif status_cell.value == "Rejected":
            status_cell.fill = red_fill
        elif status_cell.value == "Pending":
            status_cell.fill = yellow_fill

# --- Ajustar ancho de columnas ---
for column in ws.columns:
    max_length = 0
    col_letter = get_column_letter(column[0].column)
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    ws.column_dimensions[col_letter].width = max_length + 3

# --- Filtros y congelar ---
ws.auto_filter.ref = ws.dimensions
ws.freeze_panes = "A2"

# =========================
# CONTAR APPLIED, REJECTED, PENDING
# =========================
if status_col:
    applied = 0
    rejected = 0
    pending = 0
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=status_col).value
        if val == "Applied":
            applied += 1
        elif val == "Rejected":
            rejected += 1
        elif val == "Pending":
            pending += 1

    print(f"\nConteo de postulaciones {len(applications)} :")
    print(f"   Applied:  {applied}")
    print(f"   Rejected: {rejected}")
    print(f"   Pending:  {pending}")
else:
    print(" No se encontró la columna 'Status'.")
    
    
# Calcular contadores si no existen
if status_col:
    applied_count = sum(1 for row in range(2, ws.max_row + 1) if ws.cell(row=row, column=status_col).value == "Applied")
    rejected_count = sum(1 for row in range(2, ws.max_row + 1) if ws.cell(row=row, column=status_col).value == "Rejected")
    pending_count = sum(1 for row in range(2, ws.max_row + 1) if ws.cell(row=row, column=status_col).value == "Pending")
    
# =========================
# ESCRIBIR RESUMEN EN EL EXCEL
# =========================
if status_col:
    # Fila donde empieza el resumen (debajo de los datos)
    summary_row = ws.max_row + 2
    
    ws.cell(row=summary_row, column=1, value="RESUMEN")
    ws.cell(row=summary_row, column=1).font = Font(bold=True, size=12)
    
    ws.cell(row=summary_row + 1, column=1, value="Applied:")
    ws.cell(row=summary_row + 1, column=status_col, value=applied_count)
    
    ws.cell(row=summary_row + 2, column=1, value="Rejected:")
    ws.cell(row=summary_row + 2, column=status_col, value=rejected_count)
    
    ws.cell(row=summary_row + 3, column=1, value="Pending:")
    ws.cell(row=summary_row + 3, column=status_col, value=pending_count)
    
    # Negrita para los números
    for i in range(1, 4):
        ws.cell(row=summary_row + i, column=status_col).font = Font(bold=True)

# =========================
# GUARDAR
# =========================
wb.save("applications_tracker.xlsx")
print("\nArchivo actualizado correctamente.")